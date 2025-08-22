from fastapi import FastAPI
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
import os

app = FastAPI()

excel_file = "solidworks_tracker.xlsx"

# Eğer Excel yoksa başlıklarla oluştur
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Proje No", "Parça Kodu", "Parça Adı", "Durum", 
        "Tezgah Türü", "Satın Alma Durumu", "Konum",
        "Fason Firma", "Gönderim Tarihi", "Termin Tarihi", "Notlar"
    ])
    wb.save(excel_file)

class PartRecord(BaseModel):
    projeNo: str
    parcaKodu: str
    parcaAdi: str
    durum: str
    machineType: str | None = ""
    purchaseStatus: str | None = ""
    location: str | None = ""
    outsourceCompany: str | None = ""
    outsourceDate: str | None = ""
    dueDate: str | None = ""
    notes: str | None = ""

@app.post("/add_record")
def add_record(record: PartRecord):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([
        record.projeNo, record.parcaKodu, record.parcaAdi, record.durum,
        record.machineType, record.purchaseStatus, record.location,
        record.outsourceCompany, record.outsourceDate, record.dueDate, record.notes
    ])
    wb.save(excel_file)
    return {"message": "Kayıt başarıyla eklendi"}

# Tarayıcıda test için GET route
@app.get("/")
def root():
    return {"message": "Backend çalışıyor!"}
